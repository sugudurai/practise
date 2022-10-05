import sys
import datetime
from unicodedata import category
import xlsxwriter
import os
import openpyxl
import time
ticketnumber=100
class Theater:
    def __init__(self,thertr,area):
               self.theatername=thertr
               self.theaterarea=area
    def showtheaterDetails(self):
        print(  self.theatername)
        print(  self.theaterarea)
        return self.theatername+self.theaterarea
    def getTheater(self):
        return self.theatername+self.theaterarea
    
    def theatercapacity(self,row,colom):
        self.row=row
        self.coloum=colom
        return self.row,self.coloum                
        return capacity
    def createSeatNumber(self,row,coloum):
       self.rows=[]
       self.row =row
       self.coloum =coloum
       for x in range(0,self.row):
            self.columns=[]
            for y in range(0,self.coloum):
                self.columns.append('r'+str(x)+str(y))
            self.rows.append(self.columns)
       return self.rows
    def writingintoexcel(self,num1,num2,name):
        self.name=name
        result1=[]
        result=self.createSeatNumber(num1,num2)
        for x in result:
            for y in x:
               result1.append(y)
        with xlsxwriter.Workbook(self.name) as workbook:
            worksheet=workbook.add_worksheet()
            row_num=0
            for data in result1:
                worksheet.write(row_num,0,data)
                row_num=row_num+1
    def GenerateTicketNumber(self):
        global ticketnumber
        ticketnumber=ticketnumber+1
        return (ticketnumber)

    def dateTime(self):
        now=datetime.datetime.now()
        result=(now.strftime("%d-%m-%y %I:%M %p"))
        return result

    def seatnumbersheet(self,name,inpt):
        reservedrseat1=[]
        self.name=name
        self.inpt=inpt
        directory=self.name
        parent_dir="C:/Users/ADMIN/Desktop/newproject"
        path=os.path.join(parent_dir,directory)
        wb_obj=openpyxl.load_workbook(path)
        sheet_obj=wb_obj.active
        worksheet=wb_obj['Sheet1']
        row=sheet_obj.max_row
        column=sheet_obj.max_column
        reservedrseat=0
        reservedrseat=reservedrseat+self.inpt
        cell=sheet_obj.cell(1,1)
        if (cell.value==None):
            return False
        
        if self.inpt>row:
            print("YOU HAVE ENTER GREATER THAN AVAILABE SEATS: please enter within avvailabe seats  ", row)
            return False
        else:
        
         for i in range (1,row+1):
            for j in range(1,reservedrseat+1):
             cell_obj=sheet_obj.cell(row=j,column=i)
             if cell_obj.value!=None:
              reservedrseat1.append(cell_obj.value)
              #print(cell_obj.value)
        sheet_obj.delete_rows(1,reservedrseat)         
        wb_obj.save(path)
        return [reservedrseat1,True]


        
class Movie:
    item=[]
    def __init__(self,name):
        #self.item=[]
        self.moviename=name
        #self.movietime=time
        self.item.append(self.moviename)
    def movietiming(self):
        self.timing=[]
        time1=datetime.time(9,30)
        time2=datetime.time(12,30)
        time3=datetime.time(15,30)
        t1=time1.strftime('%H:%M %p')
        t2=time2.strftime('%H:%M %p')
        t3=time3.strftime('%H:%M %p')
        self.timing.append(t1)
        self.timing.append(t2)
        self.timing.append(t3)
        return self.timing
    def getMovie(self,inpu):
        self.inpu=inpu-1
        return (self.item[self.inpu]+ "  " +self.timing[self.inpu])
    def getMovie1(self,inpu):
        self.inpu=inpu-1
        return self.item[0]+self.timing[self.inpu]
    def timeComparison(self,input):
        self.input=input-1
        movie_timing=self.timing[self.input]
        my_date_time=datetime.datetime.strptime(movie_timing,"%H:%M %p")
        old_time=my_date_time.strftime("%H:%M %p")
        current_time=datetime.datetime.now()
        now_timing=current_time.strftime('%H:%M %p')
        if old_time <now_timing: 
          return True
        else:
            return False
    

    def showMoviedetails(self):
      self. res= self.movietiming()
      for x,tim in zip(self.item,self.res):
        print(x, tim)
   
    def repetedMovie(self):
       # self.res= random.randrange(0,3)
        self.item1=self.movietiming()
        self.res1=self.item[0]
        for x in self.item1:
           print(self.res1+ x)
    def showmovieRate(self):
        self.classwise=("I-class","II-class","III-class")
        self.rate=[400,350,300]
        for x,y in zip(self.classwise,self.rate):
            print(x,y)
    def getMovieRate(self,inp):
        self.inp=inp-1
        print (self.rate[self.inp])
        return (self.rate[self.inp])

def main():

 theater1=Theater("  sathyam","   anna nagar")
 theater2=Theater("2.lux mini","   purasai")
 movie1=Movie("kabhali")
 movie2=Movie("vikram")
 movie3=Movie("dagadi")
 theater1.writingintoexcel(10,5,"test6.xlsx")
 theater1.writingintoexcel(10,5,"test7.xlsx")
 theater1.writingintoexcel(10,5,"test8.xlsx")
 theater2.writingintoexcel(5,5,"test10.xlsx")
 theater2.writingintoexcel(5,5,"test11.xlsx")
 theater2.writingintoexcel(5,5,"test12.xlsx")
 done=False

 while done==False:
            print("""===========THERATER MANAGEMENT================
                          1.MOVIE DETAILS IN SATHYAM
                          2.MOVIE DETAILS IN LUXE MINI
                          3.EXIT""")
            choice=int(input("Enter choice:"))
            if choice==1:
                 movie1.showMoviedetails()
                 movie1.movietiming()
                 userinput1=int(input("Enter your choice from 1-3:"))
                 if userinput1==1:
                    res=movie1.getMovie(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet1=  theater1.getTheater()
                    print(res +" "+  theaterdet1)
                    timecompare= movie1.timeComparison(userinput1)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want"))
                     seatcount=theater1.seatnumbersheet("test6.xlsx",userinpt)
                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:
                      
                      DateTime=  theater1.dateTime()
                      movie1.showmovieRate()

                      category1=int(input("enter which class you want want"))
                      print(category1)
                      rate= movie1.getMovieRate(category1)
                      total=userinpt*rate
                      Ticketnumber=theater1.GenerateTicketNumber()
                      print("TicketNumber :",Ticketnumber) 
                      print("MOVIE NAME AND TIMING:", res+  theaterdet1)
                      print("SEATNUMBERS",seatcount[0])
                      print("DATE",DateTime)
                      print("The total amount is",total)
                 elif userinput1==2:

                    res=movie1.getMovie(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet1=  theater1.getTheater()
                    print(res+  theaterdet1)
                    timecompare= movie1.timeComparison(userinput1)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want"))
                     seatcount=theater1.seatnumbersheet("test7.xlsx",userinpt)
                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:
                      
                      DateTime=  theater1.dateTime()
                      movie1.showmovieRate()

                      category1=int(input("enter which class you want want"))
                      print(category1)
                      rate= movie1.getMovieRate(category1)
                      #print(rate)
                      total=userinpt*rate
                      Ticketnumber=theater1.GenerateTicketNumber()
                      print("TicketNumber :",Ticketnumber) 
                      print("MOVIE NAME AND TIMING:", res+  theaterdet1)
                      print("SEATNUMBERS",seatcount[0])
                      print("DATE",DateTime)
                      print("The total amount is",total)


                 elif userinput1==3:

                    res=movie1.getMovie(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet1=  theater1.getTheater()
                    print(res+" " +theaterdet1)
                    timecompare= movie1.timeComparison(userinput1)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want"))
                     seatcount=theater2.seatnumbersheet("test8.xlsx",userinpt)
                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:

                      
                       DateTime=  theater2.dateTime()
                       movie1.showmovieRate()

                       category1=int(input("enter which class you want want"))
                       print(category1)
                       rate= movie1.getMovieRate(category1)
                      #print(rate)
                       total=userinpt*rate
                       Ticketnumber=theater1.GenerateTicketNumber()
                       print("TicketNumber :",Ticketnumber) 
                       print("MOVIE NAME AND TIMING:", res+" "  +theaterdet1)
                       print("SEATNUMBERS",seatcount[0])
                       print("DATE",DateTime)
                       print("The total amount is",total)
            elif choice==2:
             movie1.repetedMovie()
             movie1.movietiming()
             userinput1=int(input("Enter your choice from 1-3:"))
             if userinput1==1:
                    res=movie1.getMovie1(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet2=  theater2.getTheater()
                    print(res +" "+  theaterdet2)
                    timecompare= movie1.timeComparison(userinput1)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want"))
                     seatcount=theater2.seatnumbersheet("test10.xlsx",userinpt)
                     
                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:
                      
                      DateTime=  theater2.dateTime()
                      movie1.showmovieRate()

                      category1=int(input("enter which class you want want: "))
                      print(category1)
                      rate= movie1.getMovieRate(category1)
                      #print(rate)
                      total=userinpt*rate
                      Ticketnumber=theater2.GenerateTicketNumber()
                      print("TicketNumber :",Ticketnumber) 
                      print("MOVIE NAME AND TIMING:", res+  theaterdet1)
                      print("SEATNUMBERS",seatcount[0])
                      print("DATE",DateTime)
                      print("The total amount is",total)
                    
             elif userinput1==2:

                    res=movie1.getMovie1(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet2=  theater2.getTheater()
                    print(res +" "+  theaterdet2)
                    timecompare= movie1.timeComparison(userinput1)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want: "))
                     seatcount=theater2.seatnumbersheet("test11.xlsx",userinpt)
                     
                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:
                      
                      DateTime=  theater2.dateTime()
                      movie1.showmovieRate()

                      category1=int(input("enter which class you want want: "))
                      print(category1)
                      rate= movie1.getMovieRate(category1)
                      #print(rate)
                      total=userinpt*rate
                      Ticketnumber=theater2.GenerateTicketNumber()
                      print("TicketNumber :",Ticketnumber) 
                      print("MOVIE NAME AND TIMING:", res+" " +theaterdet2)
                      print("SEATNUMBERS",seatcount[0])
                      print("DATE",DateTime)
                      print("The total amount is",total)


             elif userinput1==3:

                    res=movie1.getMovie1(userinput1)
                    timecompare= movie1.timeComparison(userinput1)
                    theaterdet2=  theater2.getTheater()
                    print(res+" "  +theaterdet2)
                    if(timecompare==True):
                        print("The counter is closed")
                    else:
                    
                     userinpt=int(input("enter how many tickets you want: "))
                     seatcount=theater2.seatnumbersheet("test12.xlsx",userinpt)

                     if seatcount  ==False:
                        print("NO SEATS AVAILABE")
                     else:
                      
                      DateTime=  theater2.dateTime()
                      movie1.showmovieRate()

                      category1=int(input("enter which class you want want: "))
                      print(category1)
                      rate= movie1.getMovieRate(category1)
                      #print(rate)
                      total=userinpt*rate
                      Ticketnumber=theater2.GenerateTicketNumber()
                      print("TicketNumber :",Ticketnumber) 
                      print("MOVIE NAME AND TIMING:", res+" "  +theaterdet2)
                      print("SEATNUMBERS",seatcount[0])
                      print("DATE",DateTime)
                      print("The total amount is",total)
                    
            else:
               sys.exit()
main()